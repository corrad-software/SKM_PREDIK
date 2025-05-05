import pandas as pd
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_ledger_sheet(wb, data):
    # Create ledger sheet
    ws = wb.create_sheet("Ledger")
    
    # Define styles
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center')
    right_alignment = Alignment(horizontal='right')
    yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    
    # Write initial headers
    ws.cell(row=1, column=1, value='JENIS AKAUN').font = header_font
    ws.cell(row=1, column=2, value='BUTIRAN').font = header_font
    ws.cell(row=1, column=3, value='PIC').font = header_font

    # Get the columns from LEDGER.json
    columns = data['ledger']['columns']
    current_col = 4  # Start after PIC column

    # Write column headers and merge cells for each main column
    for column in columns:
        start_col = current_col
        end_col = current_col + 1  # Each main column has 2 sub-columns (DEBIT RM, KREDIT RM)
        
        # Write and merge main column header
        ws.cell(row=1, column=start_col, value=column['name']).font = header_font
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        
        # Write sub-columns
        ws.cell(row=2, column=start_col, value=column['subColumns'][0]).font = header_font  # DEBIT RM
        ws.cell(row=2, column=end_col, value=column['subColumns'][1]).font = header_font    # KREDIT RM
        
        current_col += 2

    # Apply borders and alignment to headers
    for row in range(1, 3):
        for col in range(1, current_col):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = center_alignment

    # Start writing data from row 3
    current_row = 3

    # Write sections and their items
    for section in data['ledger']['sections']:
        # Write section header with yellow background
        ws.cell(row=current_row, column=2, value=section['name']).fill = yellow_fill
        
        # Initialize all value cells for the section row with 0.00
        col_index = 4
        for _ in columns:
            ws.cell(row=current_row, column=col_index, value='0.00').number_format = '#,##0.00'
            ws.cell(row=current_row, column=col_index+1, value='0.00').number_format = '#,##0.00'
            col_index += 2
        
        current_row += 1

        # Write groups and their items
        for group in section['groups']:
            # Only write the group header (not the items) if it's a header type
            if group.get('type') == 'header':
                ws.cell(row=current_row, column=1, value=group['code'])
                ws.cell(row=current_row, column=2, value=group['name']).fill = yellow_fill
                
                # Initialize all value cells for the group header row with 0.00
                col_index = 4
                for _ in columns:
                    ws.cell(row=current_row, column=col_index, value='0.00').number_format = '#,##0.00'
                    ws.cell(row=current_row, column=col_index+1, value='0.00').number_format = '#,##0.00'
                    col_index += 2
                
                current_row += 1
            
            # Write items for the group
            for item in group.get('items', []):
                # Write item code and name
                ws.cell(row=current_row, column=1, value=item['code'])
                ws.cell(row=current_row, column=2, value=item['name'])
                
                # Write values for each column
                col_index = 4
                for column in columns:
                    col_name = column['name']
                    values = item.get('values', {}).get(col_name, {})
                    
                    debit_cell = ws.cell(row=current_row, column=col_index, value=values.get('debit', '0.00'))
                    kredit_cell = ws.cell(row=current_row, column=col_index+1, value=values.get('kredit', '0.00'))
                    
                    debit_cell.number_format = '#,##0.00'
                    kredit_cell.number_format = '#,##0.00'
                    debit_cell.alignment = right_alignment
                    kredit_cell.alignment = right_alignment
                    
                    col_index += 2
                
                current_row += 1
            
            # Write grand total if it's a grand total type
            if group.get('type') == 'grandTotal':
                ws.cell(row=current_row, column=1, value=group['code'])
                ws.cell(row=current_row, column=2, value=group['name']).font = Font(bold=True)
                
                # Write the actual values for the grand total
                col_index = 4
                for column in columns:
                    col_name = column['name']
                    values = group.get('values', {}).get(col_name, {})
                    
                    debit_cell = ws.cell(row=current_row, column=col_index, value=values.get('debit', '0.00'))
                    kredit_cell = ws.cell(row=current_row, column=col_index+1, value=values.get('kredit', '0.00'))
                    
                    debit_cell.number_format = '#,##0.00'
                    kredit_cell.number_format = '#,##0.00'
                    debit_cell.alignment = right_alignment
                    kredit_cell.alignment = right_alignment
                    
                    col_index += 2
                
                current_row += 1

    # Apply borders to all cells
    for row in range(1, current_row):
        for col in range(1, current_col):
            ws.cell(row=row, column=col).border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 15  # JENIS AKAUN
    ws.column_dimensions['B'].width = 40  # BUTIRAN
    ws.column_dimensions['C'].width = 15  # PIC
    
    # Set width for all value columns
    for col in range(4, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_materiality_sheet(wb, data):
    ws = wb.create_sheet("Materiality")
    
    # Headers
    headers = ['Category', 'Level', 'Description', 'Recommendations']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Data
    current_row = 2
    for category in data['materiality']['categories']:
        ws.cell(row=current_row, column=1, value=category['name'])
        ws.cell(row=current_row, column=2, value=category['level'])
        ws.cell(row=current_row, column=3, value=category['description'])
        ws.cell(row=current_row, column=4, value='\n'.join(category['recommendations']))
        current_row += 1
    
    # Overall Level
    ws.cell(row=current_row + 1, column=1, value='Overall Level')
    ws.cell(row=current_row + 1, column=2, value=data['materiality']['overallLevel'])

def create_audit_sampling_sheet(wb, data):
    ws = wb.create_sheet("Audit Sampling")
    
    # Headers
    headers = ['Category', 'Level', 'Description', 'Recommendations']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Data
    current_row = 2
    for category in data['auditSampling']['categories']:
        ws.cell(row=current_row, column=1, value=category['name'])
        ws.cell(row=current_row, column=2, value=category['level'])
        ws.cell(row=current_row, column=3, value=category['description'])
        ws.cell(row=current_row, column=4, value='\n'.join(category['recommendations']))
        current_row += 1
    
    # Overall Level
    ws.cell(row=current_row + 1, column=1, value='Overall Level')
    ws.cell(row=current_row + 1, column=2, value=data['auditSampling']['overallLevel'])

def create_risk_assessment_sheet(wb, data):
    ws = wb.create_sheet("Risk Assessment")
    
    # Headers
    headers = ['Category', 'Level', 'Status', 'Description', 'Recommendations']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Data
    current_row = 2
    for category in data['riskAssessment']['categories']:
        ws.cell(row=current_row, column=1, value=category['name'])
        ws.cell(row=current_row, column=2, value=category['level'])
        ws.cell(row=current_row, column=3, value=category['status'])
        ws.cell(row=current_row, column=4, value=category['description'])
        ws.cell(row=current_row, column=5, value='\n'.join(category['recommendations']))
        current_row += 1
    
    # Overall Risk
    ws.cell(row=current_row + 1, column=1, value='Overall Risk')
    ws.cell(row=current_row + 1, column=2, value=data['riskAssessment']['overallRisk'])

def main():
    # Get the directory of this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Paths for input and output
    json_path = os.path.join(os.path.dirname(script_dir), './LEDGER.json')
    output_path = os.path.join(script_dir, 'financial_report.xlsx')
    
    # Read JSON file
    with open(json_path, 'r') as file:
        data = json.load(file)
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    create_ledger_sheet(wb, data)
    create_materiality_sheet(wb, data)
    create_audit_sampling_sheet(wb, data)
    create_risk_assessment_sheet(wb, data)
    
    # Auto-adjust column widths
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel file generated successfully at: {output_path}")

if __name__ == "__main__":
    main() 